"""Havant Thicket Material Testing App - Data processing module."""
#import os
import re
from datetime import datetime
import numpy as np
import pandas as pd
import streamlit as st
from io import BytesIO
import boto3
from botocore.exceptions import ClientError

# Function to highlight rows based on Status column
def highlight_status(row):
    if 'Status' in row.index:
        status = row['Status']
        if status == 'FAIL':
            return ['background-color: #ffcccc; color: black'] * len(row)  # Light red
        if status == 'CAUTION':
            return ['background-color: #ffffcc; color: black'] * len(row)  # Light yellow
        if status == 'PASS':
            return [''] * len(row)
    return [''] * len(row)

st.set_page_config(layout="wide")

#with st.sidebar:
#    st.title("Navigation")
#    st.markdown("---")

st.title('Havant Thicket Material Testing App')

uploaded_files = st.file_uploader(
    "Upload Excel file(s) of Daily Published Results:",
    type=['xlsx', 'xls'],
    accept_multiple_files=True
)

# Upload files to S3
def upload_to_s3(file_name, file_bytes):
    """Upload file to S3 bucket if it doesn't already exist."""
    s3_client = boto3.client('s3', region_name='us-west-2')
    bucket_name = 'htr-qa-bucket-001'
    s3_key = f'Summary_Sheets/{file_name}'
    
    try:
        # Check if file already exists in S3
        s3_client.head_object(Bucket=bucket_name, Key=s3_key)
        st.warning(f"File '{file_name}' already exists in S3. See Table Tab for details.")
        return False
    except ClientError as e:
        if e.response['Error']['Code'] == '404':
            # File doesn't exist, proceed with upload
            try:
                s3_client.upload_fileobj(
                    BytesIO(file_bytes),
                    bucket_name,
                    s3_key
                )
                st.success(f"File '{file_name}' uploaded to S3 successfully.")
                return True
            except ClientError as upload_error:
                st.error(f"Error uploading file to S3: {upload_error}")
                return False
        else:
            st.error(f"Error checking S3 file: {e}")
            return False


# Process uploaded files and upload to S3
dataframes = {}
if uploaded_files:
    # Store file bytes once so the Streamlit UploadedFile is not reused after closing
    uploaded_file_bytes = {
        uploaded_file.name: uploaded_file.getvalue()
        for uploaded_file in uploaded_files
    }

    for uploaded_file in uploaded_files:
        upload_to_s3(
            uploaded_file.name,
            uploaded_file_bytes[uploaded_file.name]
        )

    all_field_data = []
    all_lab_data = []

    for file_name, file_bytes in uploaded_file_bytes.items():
        excel_data = pd.ExcelFile(BytesIO(file_bytes))
        sheets_to_import = [
            sheet for sheet in [
                'Field', 'Lab', 'Field ', 'Lab ',
                ' Field', ' Lab'
            ]
            if sheet in excel_data.sheet_names
        ]
        
        for sheet in sheets_to_import:
            sheet_data = excel_data.parse(sheet)
            sheet_key = sheet.replace(" ", "")
            
            if sheet_key == "Field":
                all_field_data.append(sheet_data)
            elif sheet_key == "Lab":
                all_lab_data.append(sheet_data)
    
    # Amalgamate and remove duplicates
    if all_field_data:
        field_combined = pd.concat(
            all_field_data,
            ignore_index=True
        )
        sample_id_col = None
        if 'Sample ID' in field_combined.columns:
            sample_id_col = 'Sample ID'
        elif 'SOCOTE Sample Reference' in field_combined.columns:
            sample_id_col = 'SOCOTE Sample Reference'
        
        if sample_id_col:
            field_combined = (
                field_combined.drop_duplicates(
                    subset=[sample_id_col],
                    keep='first'
                )
            )
        dataframes["Field"] = field_combined
    
    if all_lab_data:
        lab_combined = pd.concat(
            all_lab_data,
            ignore_index=True
        )
        sample_id_col = None
        if 'Sample ID' in lab_combined.columns:
            sample_id_col = 'Sample ID'
        elif 'SOCOTE Sample Reference' in lab_combined.columns:
            sample_id_col = 'SOCOTE Sample Reference'
        
        if sample_id_col:
            lab_combined = (
                lab_combined.drop_duplicates(
                    subset=[sample_id_col],
                    keep='first'
                )
            )
        dataframes["Lab"] = lab_combined
    
# Remove parentheses and their contents from all column names in Field and Lab dataframes
for key in ["Field", "Lab"]:
    if key in dataframes and not dataframes[key].empty:
        dataframes[key].columns = [
            re.sub(r"\s*\([^)]*\)", "", col).strip().rstrip() for col in dataframes[key].columns
        ]
        
# Create a new dataframe for each property
CORE_QUT = pd.DataFrame()
SHOULDER_QUT = pd.DataFrame()
Liquid_Limits = pd.DataFrame()
Plasticity_Index = pd.DataFrame()
CORE_HSV = pd.DataFrame()
SHOULDER_HSV = pd.DataFrame()
FORMATION_HSV = pd.DataFrame()
CORE_NDG = pd.DataFrame()
SHOULDER_NDG = pd.DataFrame()
SAND_NDG = pd.DataFrame()
CORE_SRT = pd.DataFrame()
SHOULDER_SRT = pd.DataFrame()
SAND_PSD = pd.DataFrame()
STONE_PSD = pd.DataFrame()

##------------------------------------------ FIELD TESTS --------------------------------------------------##

#Core QUT Frame
if "Lab" in dataframes and 'Triaxial Shear Strength' in dataframes["Lab"].columns:
    mask = (
        dataframes["Lab"]['Triaxial Shear Strength'].notna() &
        dataframes["Lab"]['General Location'].str.contains('core', case=False, na=False)
    )
    CORE_QUT = dataframes["Lab"][mask].copy()
    cols_to_keep = list(CORE_QUT.columns[:6]) + ['Triaxial Shear Strength'] + ['Specification']
    CORE_QUT = CORE_QUT[[col for col in cols_to_keep if col in CORE_QUT.columns]]

    # Remove 'Date Reported'
    CORE_QUT = CORE_QUT.drop(columns=['Date Reported'], errors='ignore')
    
    CORE_QUT['Status'] = CORE_QUT.iloc[:, 5].apply(lambda x: 'PASS' if pd.notna(x) and 55 <= float(x) <= 110 else ('CAUTION' if pd.notna(x) and 50 <= float(x) <= 120 else 'FAIL') if isinstance(x, (int, float)) or (isinstance(x, str) and x.replace('.','',1).isdigit()) else 'FAIL')

#Shoulder QUT Frame
if "Lab" in dataframes and 'Triaxial Shear Strength' in dataframes["Lab"].columns:
    mask = (
        dataframes["Lab"]['Triaxial Shear Strength'].notna() &
        dataframes["Lab"]['General Location'].str.contains('shoulder', case=False, na=False)
    )
    SHOULDER_QUT = dataframes["Lab"][mask].copy()
    cols_to_keep = list(SHOULDER_QUT.columns[:6]) + ['Triaxial Shear Strength'] + ['Specification']
    SHOULDER_QUT = SHOULDER_QUT[[col for col in cols_to_keep if col in SHOULDER_QUT.columns]]

    # Remove 'Date Reported'
    SHOULDER_QUT = SHOULDER_QUT.drop(columns=['Date Reported'], errors='ignore')
    
    SHOULDER_QUT['Status'] = SHOULDER_QUT.iloc[:, 5].apply(lambda x: 'PASS' if pd.notna(x) and 65 <= float(x) <= 200 else 'FAIL')
    
#Core HSV Frame
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('Hand Shear Vane', case=False, na=False) &
        dataframes["Field"]['Location'].str.contains('core', case=False, na=False) &
        ~dataframes["Field"]['Location'].str.contains('formation', case=False, na=False) &
        ~dataframes["Field"]['Location'].str.contains(r'formation|sub-formation', case=False, na=False, regex=True)
    )
    CORE_HSV = dataframes["Field"][mask].copy()
    cols_to_keep = list(CORE_HSV.columns[:8]) + ['Hand Vane Shear Strength']
    CORE_HSV = CORE_HSV[[col for col in cols_to_keep if col in CORE_HSV.columns]]
    
    # Remove columns 2, 3, 5 (indices)
    cols_to_drop = [CORE_HSV.columns[i] for i in [1, 2, 3, 5] if i < len(CORE_HSV.columns)]
    CORE_HSV = CORE_HSV.drop(columns=cols_to_drop, errors='ignore')
    
    # Remove operators from column 4
    col = CORE_HSV.columns[4]
    CORE_HSV[col] = (CORE_HSV[col].astype(str).str.replace(r"[<>=]", "", regex=True).str.strip())   
    CORE_HSV[col] = pd.to_numeric(CORE_HSV[col], errors="coerce")
    
    # Add status column based on column 4 value
    CORE_HSV['Status'] = CORE_HSV.iloc[:, 4].apply(lambda x: 'PASS' if pd.notna(x) and 55 <= float(x) <= 110 else ('CAUTION' if pd.notna(x) and 50 <= float(x) <= 120 else 'FAIL'))

#Shoulder HSV Frame    
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('Hand Shear Vane', case=False, na=False) &
        dataframes["Field"]['Location'].str.contains('shoulder', case=False, na=False) &
        ~dataframes["Field"]['Location'].str.contains(r'formation|sub-formation', case=False, na=False, regex=True)
    )
    SHOULDER_HSV = dataframes["Field"][mask].copy()
    cols_to_keep = list(SHOULDER_HSV.columns[:8]) + ['Hand Vane Shear Strength']
    SHOULDER_HSV = SHOULDER_HSV[[col for col in cols_to_keep if col in SHOULDER_HSV.columns]]
    
    # Remove columns 1, 2, 3, 5 (indices)
    cols_to_drop = [SHOULDER_HSV.columns[i] for i in [1, 2, 3, 5] if i < len(SHOULDER_HSV.columns)]
    SHOULDER_HSV = SHOULDER_HSV.drop(columns=cols_to_drop, errors='ignore')
    
    # Remove operators from column 4
    col = SHOULDER_HSV.columns[4]
    SHOULDER_HSV[col] = (SHOULDER_HSV[col].astype(str).str.replace(r"[<>=]", "", regex=True).str.strip())   
    SHOULDER_HSV[col] = pd.to_numeric(SHOULDER_HSV[col], errors="coerce")
    
    # Add status column based on column 4 value
    SHOULDER_HSV['Status'] = SHOULDER_HSV.iloc[:, 4].apply(lambda x: 'PASS' if pd.notna(x) and 65 <= float(x) <= 200 else 'FAIL')
    
#Formation HSV Frame
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('Hand Shear Vane', case=False, na=False) &
        dataframes["Field"]['Location'].str.contains('formation', case=False, na=False) &
        ~dataframes["Field"]['Location'].str.contains(r'formation.*sub|sub.*formation', case=False, na=False, regex=True)
    )
    FORMATION_HSV = dataframes["Field"][mask].copy()
    cols_to_keep = list(FORMATION_HSV.columns[:8]) + ['Hand Vane Shear Strength']
    FORMATION_HSV = FORMATION_HSV[[col for col in cols_to_keep if col in FORMATION_HSV.columns]]
    
    # Remove columns 1, 2, 3, 5 (indices)
    cols_to_drop = [FORMATION_HSV.columns[i] for i in [1, 2, 3, 5] if i < len(FORMATION_HSV.columns)]
    FORMATION_HSV = FORMATION_HSV.drop(columns=cols_to_drop, errors='ignore')
    
    # Remove operators from column 4
    FORMATION_HSV.iloc[:, 4] = FORMATION_HSV.iloc[:, 4].astype(str).str.replace(r'[<>=]', '', regex=True)
    
    # Add status column based on column 4 value
    FORMATION_HSV['Status'] = FORMATION_HSV.iloc[:, 4].apply(lambda x: 'PASS' if pd.notna(x) and float(x) >= 50 else 'FAIL')

#CORE NDG Frame    
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('In-Situ Density by Nuclear Method', case=False, na=False) & 
        dataframes["Field"]['Location'].str.contains('core', case=False, na=False) &
        ~dataframes["Field"].iloc[:, 6].str.contains('sand|0-4|filter', case=False, na=False, regex=True) &
        ~dataframes["Field"].iloc[:, 7].str.contains('sand|0-4|filter|drain', case=False, na=False)
    )
    CORE_NDG = dataframes["Field"][mask].copy() 

    CORE_NDG = CORE_NDG.replace('', np.nan)
    CORE_NDG = CORE_NDG.dropna(axis=1, how='all')
    
    # Remove columns 1, 2, 3, 5 (indices)
    cols_to_drop = [CORE_NDG.columns[i] for i in [1, 2, 3, 5] if i < len(CORE_NDG.columns)]
    CORE_NDG = CORE_NDG.drop(columns=cols_to_drop, errors='ignore')
    
    # Rename column 5 to 'Degree of Compaction'
    if len(CORE_NDG.columns) > 5:
        CORE_NDG.columns.values[5] = 'Degree of Compaction'
    
    # Add status column based on column 5 value
    if len(CORE_NDG.columns) > 6:
        CORE_NDG['Status'] = (
            CORE_NDG.iloc[:, 6].apply(
                lambda x: 'PASS' if pd.notna(x) and (
                    isinstance(x, (int, float)) or 
                    pd.to_numeric(x, errors='coerce') is not None
                ) and float(x) <= 5 else 'FAIL'
            )
        )
      
#SHOULDER NDG Frame    
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('In-Situ Density by Nuclear Method', case=False, na=False) & 
        dataframes["Field"]['Location'].str.contains('shoulder', case=False, na=False) &
        ~dataframes["Field"].iloc[:, 6].str.contains('sand|0-4|filter', case=False, na=False, regex=True) &
        ~dataframes["Field"].iloc[:, 7].str.contains('sand|0-4|filter|drain', case=False, na=False)
    )
    SHOULDER_NDG = dataframes["Field"][mask].copy()
    
    SHOULDER_NDG = SHOULDER_NDG.replace('', np.nan)
    SHOULDER_NDG = SHOULDER_NDG.dropna(axis=1, how='all')
    
    # Remove columns 1, 2, 3, 5 (indices)
    cols_to_drop = [SHOULDER_NDG.columns[i] for i in [1, 2, 3, 5] if i < len(SHOULDER_NDG.columns)]
    SHOULDER_NDG = SHOULDER_NDG.drop(columns=cols_to_drop, errors='ignore')
    
    # Rename column 5 to 'Degree of Compaction'
    if len(SHOULDER_NDG.columns) > 5:
        SHOULDER_NDG.columns.values[5] = 'Degree of Compaction'
        
    SHOULDER_NDG['Status'] = SHOULDER_NDG.iloc[:, 5].apply(lambda x: 'PASS' if pd.notna(x) and float(x) >= 95 else 'FAIL')

#SAND NDG Frame    
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('In-Situ Density by Nuclear Method', case=False, na=False) & 
        dataframes["Field"]['Material'].str.contains('sand|0-4|filter', case=False, na=False)
    )
    SAND_NDG = dataframes["Field"][mask].copy() 

    if not SAND_NDG.empty:
        SAND_NDG = SAND_NDG.replace('', np.nan)
        SAND_NDG = SAND_NDG.dropna(axis=1, how='all')
    
        # Remove columns 1, 2, 3, 5 (indices)
        cols_to_drop = [SAND_NDG.columns[i] for i in [1, 2, 3, 5] if i < len(SAND_NDG.columns)]
        SAND_NDG = SAND_NDG.drop(columns=cols_to_drop, errors='ignore')
    
        # Rename column 5 to 'Degree of Compaction'
        if len(SAND_NDG.columns) > 5:
            SAND_NDG.columns.values[5] = 'Degree of Compaction'

        SAND_NDG['Status'] = SAND_NDG['Degree of Compaction'].apply(
        lambda x: 'PASS' if pd.notna(x) and float(x) >= 92 else 'FAIL'
    )
    
#CORE SRT Frame    
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('In-Situ Density by Sand Replacement', case=False, na=False) & 
        dataframes["Field"]['Location'].str.contains('core', case=False, na=False) &
        ~dataframes["Field"].iloc[:, 6].str.contains('sand|0-4|filter', case=False, na=False, regex=True) &
        ~dataframes["Field"].iloc[:, 7].str.contains('sand|0-4|filter|drain', case=False, na=False)
    )
    CORE_SRT = dataframes["Field"][mask].copy() 

    CORE_SRT = CORE_SRT.replace('', np.nan)
    CORE_SRT = CORE_SRT.dropna(axis=1, how='all')
    
    # Remove columns 1, 2, 3, 5 (indices)
    cols_to_drop = [CORE_SRT.columns[i] for i in [1, 2, 3, 5] if i < len(CORE_SRT.columns)]
    CORE_SRT = CORE_SRT.drop(columns=cols_to_drop, errors='ignore')
    
    # Rename column 5 to 'Degree of Compaction'
    if len(CORE_SRT.columns) > 5:
        CORE_SRT.columns.values[5] = 'Degree of Compaction'
       
    # Add status column based on column 5 value
    if len(CORE_SRT.columns) > 6:
        CORE_SRT['Status'] = (
            CORE_SRT.iloc[:, 6].apply(
                lambda x: 'PASS' if pd.notna(x) and (
                    isinstance(x, (int, float)) or 
                    pd.to_numeric(x, errors='coerce') is not None
                ) and float(x) <= 5 else 'FAIL'
            )
        )   
       
#SHOULDER SRT Frame    
if "Field" in dataframes:
    mask = (
        dataframes["Field"]['Test Name'].str.contains('In-Situ Density by Sand Replacement', case=False, na=False) & 
        dataframes["Field"]['Location'].str.contains('shoulder', case=False, na=False) &
        ~dataframes["Field"].iloc[:, 6].str.contains('sand|0-4|filter', case=False, na=False, regex=True) &
        ~dataframes["Field"].iloc[:, 7].str.contains('sand|0-4|filter|drain', case=False, na=False)
    )
    SHOULDER_SRT = dataframes["Field"][mask].copy()

    SHOULDER_SRT = SHOULDER_SRT.replace('', np.nan)
    SHOULDER_SRT = SHOULDER_SRT.dropna(axis=1, how='all')
    
    # Remove columns 1, 2, 3, 5 (indices)
    cols_to_drop = [SHOULDER_SRT.columns[i] for i in [1, 2, 3, 5] if i < len(SHOULDER_SRT.columns)]
    SHOULDER_SRT = SHOULDER_SRT.drop(columns=cols_to_drop, errors='ignore')
    
    # Rename column 5 to 'Degree of Compaction'
    if len(SHOULDER_SRT.columns) > 5:
        SHOULDER_SRT.columns.values[5] = 'Degree of Compaction'

    SHOULDER_SRT['Status'] = SHOULDER_SRT.iloc[:, 5].apply(lambda x: 'PASS' if pd.notna(x) and float(x) >= 95 else 'FAIL')

##------------------------------------------ LAB TESTS --------------------------------------------------##

#Sand PSD frame
if "Lab" in dataframes and 'Description' in dataframes["Lab"].columns:
    mask = (
        dataframes["Lab"]['Description'].str.contains('0-4|sand', case=False, na=False, regex=True)
    )
    SAND_PSD = dataframes["Lab"][mask].copy()
    
    # Keep first five columns and columns with 'Specification' and 'PSD failed' headings
    cols_to_keep = list(SAND_PSD.columns[:6])
    for col in SAND_PSD.columns[5:]:
        if ('Specification' in col or 'PSD Failed' in col) and 'Date Reported' not in col:
            cols_to_keep.append(col)
    SAND_PSD = SAND_PSD[cols_to_keep]
    
    # Add Status column based on 'PSD Failed' column
    psd_failed_col = next((col for col in SAND_PSD.columns if 'PSD Failed' in col), None)
    if psd_failed_col:
        SAND_PSD['Status'] = SAND_PSD[psd_failed_col].apply(lambda x: 'PASS' if x == 'No' else 'FAIL')

#Stone PSD frame
if "Lab" in dataframes and 'Description' in dataframes["Lab"].columns:
    mask = (
        dataframes["Lab"]['Description'].str.contains('4-20|gravel', case=False, na=False)
    )
    STONE_PSD = dataframes["Lab"][mask].copy()
    
    cols_to_keep = list(STONE_PSD.columns[:6])
    for col in STONE_PSD.columns[5:]:
        if ('Specification' in col or 'PSD Failed' in col) and 'Date Reported' not in col:
            cols_to_keep.append(col)
    STONE_PSD = STONE_PSD[cols_to_keep]
    
    # Add Status column based on 'PSD Failed' column
    psd_failed_col = next((col for col in STONE_PSD.columns if 'PSD Failed' in col), None)
    if psd_failed_col:
        STONE_PSD['Status'] = STONE_PSD[psd_failed_col].apply(lambda x: 'PASS' if x == 'No' else 'FAIL')

#LL Frame
if "Lab" in dataframes and 'LiquidLimit' in dataframes["Lab"].columns:
    mask = dataframes["Lab"]['LiquidLimit'] > 0
    Liquid_Limits = dataframes["Lab"][mask].copy()

    cols_to_keep = list(Liquid_Limits.columns[:6])
    for col in Liquid_Limits.columns[5:]:
        if 'LiquidLimit' in col:
            cols_to_keep.append(col)
    Liquid_Limits = Liquid_Limits[cols_to_keep]
    Liquid_Limits = Liquid_Limits.drop(columns=['Date Reported'], errors='ignore')
    
    # Add Status column based on LiquidLimit value
    ll_col = next((col for col in Liquid_Limits.columns if 'LiquidLimit' in col), None)
    if ll_col:
        Liquid_Limits['Status'] = Liquid_Limits[ll_col].apply(lambda x: 'FAIL' if x > 90 else 'PASS')

#PI Frame
if "Lab" in dataframes and 'Plasticity Index' in dataframes["Lab"].columns:
    mask = dataframes["Lab"]['Plasticity Index'] > 0
    Plasticity_Index = dataframes["Lab"][mask].copy()

    cols_to_keep = list(Plasticity_Index.columns[:6])
    for col in Plasticity_Index.columns[5:]:
        if 'Plasticity Index' in col:
            cols_to_keep.append(col)
    Plasticity_Index = Plasticity_Index[cols_to_keep]
    Plasticity_Index = Plasticity_Index.drop(columns=['Date Reported'], errors='ignore')
    
    # Add Status column based on Plasticity Index value
    pi_col = next((col for col in Plasticity_Index.columns if 'Plasticity Index' in col), None)
    if pi_col:
        Plasticity_Index['Status'] = Plasticity_Index[pi_col].apply(lambda x: 'PASS' if 20 <= x <= 65 else 'FAIL')
        
##------------------------------------------ DISPLAY RESULTS --------------------------------------------------##

tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13, tab14 = st.tabs([
    "Core QUT",
    "Shoulder QUT",
    "Liquid Limits",
    "Plasticity Index",
    "Core HSV",
    "Shoulder HSV",
    "Formation HSV",
    "Core NDG",
    "Shoulder NDG",
    "Sand NDG",
    "Core SRT",
    "Shoulder SRT",
    "Sand PSD",
    "Stone PSD"
])

def format_numeric_columns(df):
    """Format numeric columns to 2 decimal places."""
    return df.style.apply(
        highlight_status,
        axis=1
    ).format(
        {col: '{:.2f}' for col in df.select_dtypes(
            include=['number']
        ).columns}
    )

with tab1:
    if not CORE_QUT.empty:
        st.dataframe(
            format_numeric_columns(CORE_QUT),
            use_container_width=True
        )
    else:
        st.info('No Core QUT data available')

with tab2:
    if not SHOULDER_QUT.empty:
        st.dataframe(
            format_numeric_columns(SHOULDER_QUT),
            use_container_width=True
        )
    else:
        st.info('No Shoulder QUT data available')

with tab3:
    if not Liquid_Limits.empty:
        st.dataframe(
            format_numeric_columns(Liquid_Limits),
            use_container_width=True
        )
    else:
        st.info('No Liquid Limit data available')

with tab4:
    if not Plasticity_Index.empty:
        st.dataframe(
            format_numeric_columns(Plasticity_Index),
            use_container_width=True
        )
    else:
        st.info('No Plasticity Index data available')

with tab5:
    if not CORE_HSV.empty:
        st.dataframe(
            format_numeric_columns(CORE_HSV),
            use_container_width=True
        )
    else:
        st.info('No Core HSV data available')

with tab6:
    if not SHOULDER_HSV.empty:
        st.dataframe(
            format_numeric_columns(SHOULDER_HSV),
            use_container_width=True
        )
    else:
        st.info('No Shoulder HSV data available')

with tab7:
    if not FORMATION_HSV.empty:
        st.dataframe(
            format_numeric_columns(FORMATION_HSV),
            use_container_width=True
        )
    else:
        st.info('No Formation HSV data available')

with tab8:
    if not CORE_NDG.empty:
        st.dataframe(
            format_numeric_columns(CORE_NDG),
            use_container_width=True
        )
    else:
        st.info('No Core NDG data available')

with tab9:
    if not SHOULDER_NDG.empty:
        st.dataframe(
            format_numeric_columns(SHOULDER_NDG),
            use_container_width=True
        )
    else:
        st.info('No Shoulder NDG data available')

with tab10:
    if not SAND_NDG.empty:
        st.dataframe(
            format_numeric_columns(SAND_NDG),
            use_container_width=True
        )
    else:
        st.info('No Sand NDG data available')

with tab11:
    if not CORE_SRT.empty:
        st.dataframe(
            format_numeric_columns(CORE_SRT),
            use_container_width=True
        )
    else:
        st.info('No Core SRT data available')

with tab12:
    if not SHOULDER_SRT.empty:
        st.dataframe(
            format_numeric_columns(SHOULDER_SRT),
            use_container_width=True
        )
    else:
        st.info('No Shoulder SRT data available')

with tab13:
    if not SAND_PSD.empty:
        st.dataframe(
            format_numeric_columns(SAND_PSD),
            use_container_width=True
        )
    else:
        st.info('No Sand PSD data available')

with tab14:
    if not STONE_PSD.empty:
        st.dataframe(
            format_numeric_columns(STONE_PSD),
            use_container_width=True
        )
    else:
        st.info('No Stone PSD data available')

# Download Review Button
st.divider()
st.subheader('Download Review')

def create_excel_download():
    """Create Excel file with all material testing data."""
    from openpyxl.styles import PatternFill
    
    output = BytesIO()
    excel_writer = pd.ExcelWriter(output, engine='openpyxl')

    # Dictionary of all dataframes with their sheet names
    sheets_dict = {
        'Core QUT': CORE_QUT,
        'Shoulder QUT': SHOULDER_QUT,
        'Liquid Limits': Liquid_Limits,
        'Plasticity Index': Plasticity_Index,
        'Core HSV': CORE_HSV,
        'Shoulder HSV': SHOULDER_HSV,
        'Formation HSV': FORMATION_HSV,
        'Core NDG': CORE_NDG,
        'Shoulder NDG': SHOULDER_NDG,
        'Sand NDG': SAND_NDG,
        'Core SRT': CORE_SRT,
        'Shoulder SRT': SHOULDER_SRT,
        'Sand PSD': SAND_PSD,
        'Stone PSD': STONE_PSD
    }

    # Define colors for status formatting
    fail_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    caution_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    # Write each non-empty dataframe to its own sheet
    has_data = False
    for sheet_name, df in sheets_dict.items():
        if not df.empty:
            df.to_excel(excel_writer, sheet_name=sheet_name, index=False)
            has_data = True
            
            # Apply formatting based on Status column
            worksheet = excel_writer.sheets[sheet_name]
            if 'Status' in df.columns:
                for row_idx, status_value in enumerate(df['Status'], start=2):
                    if status_value == 'FAIL':
                        for col in worksheet.iter_cols(min_row=row_idx, max_row=row_idx):
                            for cell in col:
                                cell.fill = fail_fill
                    elif status_value == 'CAUTION':
                        for col in worksheet.iter_cols(min_row=row_idx, max_row=row_idx):
                            for cell in col:
                                cell.fill = caution_fill

    # Add a blank sheet if no data exists
    if not has_data:
        pd.DataFrame().to_excel(excel_writer, sheet_name='Empty', index=False)

    excel_writer.close()
    output.seek(0)
    return output.getvalue()

excel_data = create_excel_download()

st.download_button(
    label='📥 Download Review',
    data=excel_data,
    file_name=f'Material_Testing_Review_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)
